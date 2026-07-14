import io

import pytest
from openpyxl import Workbook

from services.dip_service import DIPService, HttpSessionDIPPortalClient
from services.invoice_archive_service import InvoiceArchiveService


def _xlsx_bytes(*, export=False):
    workbook = Workbook()
    sheet = workbook.active
    sheet.cell(1, 1, "VÝKUP - výrobní EAN" if export else "Detail faktury")
    sheet.cell(4, 1, "01.06.2026 - 30.06.2026")
    sheet.cell(4, 2, 0.184)
    sheet.cell(4, 3, 22.5)
    sheet.cell(4, 4, 528.81)
    sheet.cell(4, 5, 2873.97)
    if export:
        sheet.cell(4, 6, 75.0)
    header_row = 6 if export else 8
    headers = ["Datum", "Čas", "Množství", "Cena"]
    for column, value in enumerate(headers, 1):
        sheet.cell(header_row, column, value)
    sheet.cell(header_row + 1, 1, "01.06.2026")
    sheet.cell(header_row + 1, 2, "00:00 - 00:14")
    sheet.cell(header_row + 1, 3, 0.001)
    sheet.cell(header_row + 1, 4, 2.5)
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def test_dip_service_sync_persists_normalized_profile(tmp_path):
    class Client:
        def fetch_profile(self, _cfg):
            return {"meter_id": "1000000001", "consumption_ean": "859182400000000001"}, "<html>profile</html>"

    service = DIPService(tmp_path, client=Client())
    result = service.sync({"dip": {"enabled": True, "username": "user", "password": "secret"}})
    assert result["ok"] is True
    assert service.get_profile()["source"] == "dip"
    assert service.get_status({"dip": {"enabled": True, "username": "user", "password": "secret"}})["healthy"] is True
    assert service.get_status({"dip": {"enabled": True, "username": "user", "password": "secret"}})["state"] == "connected"


def test_dip_status_does_not_report_stale_error_when_disabled_or_unconfigured(tmp_path):
    service = DIPService(tmp_path)
    service.status_path.write_text('{"healthy": false, "last_error": {"code": "DIP_LOGIN_FAILED", "message": "old"}}', encoding="utf-8")
    disabled = service.get_status({"dip": {"enabled": False}})
    unconfigured = service.get_status({"dip": {"enabled": True, "username": ""}})
    assert disabled["state"] == "disabled"
    assert disabled["last_error"] is None
    assert unconfigured["state"] == "not_configured"
    assert unconfigured["last_error"] is None


def test_dip_normalization_keeps_billing_fields_without_bank_identifiers():
    client = HttpSessionDIPPortalClient()
    point = client._normalize_point(
        {
            "cislo": "0000000001", "ean": "859000000000000001", "typText": "Spotřeba", "jmenoOP": "Test User",
            "adresa": {"adresaComplete": "Test 1"}, "contactOM": {"firstName": "Test", "lastName": "User", "email": "test@example.invalid", "telephone": "+420000000000"},
            "faktaOM": {"napetovaHladina": "NN", "typMereni": "B", "casoveRezy": [{"pocetFazi": 3.0, "hlavniJistic": 25.0, "sazbaDistribuce": "D45D"}]},
            "elektromery": [{"sernr": "1000000000"}], "su": {"dodavatelTxt": "Test supplier", "listPlateb": [{"typ": "FAKTURA", "aktualniZpusobPlatby": {"text": "Převod"}, "aktualniBanka": {"iban": "SECRET"}}]},
        }, [], {"signals": []}, {"notifSettings": {}, "shutdowns": []}, {},
    )
    assert point["technical"]["distribution_tariff"] == "D45D"
    assert point["contract"]["payment_methods"] == [{"type": "FAKTURA", "method": "Převod"}]
    assert "SECRET" not in str(point)


def test_invoice_archive_stores_lists_audits_and_deletes_xlsx(tmp_path):
    service = InvoiceArchiveService(tmp_path)
    record = service.store("detail.xlsx", _xlsx_bytes())
    assert record["parsed"]["document_type"] == "supply_detail_xlsx"
    assert record["parsed"]["rows"][0]["Čas"] == "00:00 - 00:14"
    assert "rows" not in service.list_documents()[0]["parsed"]

    result = service.audit(record["id"], {
        "actual": {"kwh_total": 184.0},
        "invoice": {"actual": {"commercial": {"spot_energy": 530.0}}},
    })
    assert result["overall"] == "error"
    assert service.delete(record["id"]) is True
    assert service.get_document(record["id"]) is None


@pytest.mark.parametrize("filename,data", [("fake.pdf", b"not pdf"), ("fake.xlsx", b"not xlsx")])
def test_invoice_archive_rejects_spoofed_extensions(tmp_path, filename, data):
    service = InvoiceArchiveService(tmp_path)
    with pytest.raises(ValueError):
        service.store(filename, data)
    assert list(service.originals_dir.iterdir()) == []
