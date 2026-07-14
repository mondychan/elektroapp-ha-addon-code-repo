from __future__ import annotations

from datetime import datetime, timezone
import hashlib
import io
import json
from pathlib import Path
import re
import shutil
import zipfile
from typing import Any

from openpyxl import load_workbook
from pypdf import PdfReader


ALLOWED_EXTENSIONS = {".pdf", ".xlsx"}
MAX_DOCUMENT_SIZE = 25 * 1024 * 1024
MAX_XLSX_UNCOMPRESSED_SIZE = 100 * 1024 * 1024


def _number(value: str | None) -> float | None:
    if not value:
        return None
    cleaned = value.replace("\xa0", " ").replace(" ", "").replace("Kč", "").replace(",", ".")
    try:
        return float(cleaned)
    except ValueError:
        return None


def parse_invoice_pdf(data: bytes) -> dict[str, Any]:
    reader = PdfReader(io.BytesIO(data))
    text = "\n".join(page.extract_text() or "" for page in reader.pages)
    period = re.search(r"Období:\s*(\d{1,2}\.\s*\d{1,2}\.\s*\d{4})\s*-\s*(\d{1,2}\.\s*\d{1,2}\.\s*\d{4})", text)
    invoice_number = re.search(r"vyúčtování elektřiny č\.\s*(\d+)", text, re.I)
    eans = list(dict.fromkeys(re.findall(r"\b859\d{15}\b", text)))

    def amount(label: str, occurrence: int = 0) -> float | None:
        matches = re.findall(rf"{label}\s+(-?[\d\s]+,\d{{2}})\s*Kč", text, re.I)
        return _number(matches[occurrence]) if len(matches) > occurrence else None

    return {
        "document_type": "invoice_pdf",
        "invoice_number": invoice_number.group(1) if invoice_number else None,
        "period_from": datetime.strptime(period.group(1).replace(" ", ""), "%d.%m.%Y").date().isoformat() if period else None,
        "period_to": datetime.strptime(period.group(2).replace(" ", ""), "%d.%m.%Y").date().isoformat() if period else None,
        "consumption_ean": eans[0] if eans else None,
        "supply_without_vat": amount("Dodávka energií"),
        "commercial_without_vat": amount("Obchodní platby"),
        "regulated_without_vat": amount("Regulované platby"),
        "advance_with_vat": abs(amount("Vyúčtované zálohy", 1) or 0.0) or None,
        "export_total": abs(amount("Výkup elektřiny") or 0.0) or None,
        "settlement": amount("Výsledek vyúčtování"),
        "raw_text": text,
    }


def parse_invoice_xlsx(data: bytes) -> dict[str, Any]:
    workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    title = str(sheet.cell(1, 1).value or "")
    is_export = "VÝKUP" in title.upper() or "výrobní EAN" in title
    summary_row = [sheet.cell(4, col).value for col in range(1, 7)]
    period_match = re.match(r"\s*(\d{2}\.\d{2}\.\d{4})\s*-\s*(\d{2}\.\d{2}\.\d{4})\s*", str(summary_row[0] or ""))
    header_row = 6 if is_export else 8
    headers = [sheet.cell(header_row, col).value for col in range(1, 11 if is_export else 9)]
    rows = []
    for row in sheet.iter_rows(min_row=header_row + 1, max_col=len(headers), values_only=True):
        if not row[0] or not row[1]:
            continue
        rows.append(dict(zip(headers, row)))
    return {
        "document_type": "export_detail_xlsx" if is_export else "supply_detail_xlsx",
        "title": title,
        "period": summary_row[0],
        "period_from": datetime.strptime(period_match.group(1), "%d.%m.%Y").date().isoformat() if period_match else None,
        "period_to": datetime.strptime(period_match.group(2), "%d.%m.%Y").date().isoformat() if period_match else None,
        "quantity_mwh": summary_row[1],
        "total_eur": summary_row[2],
        "total_czk": summary_row[3],
        "average_czk_mwh": summary_row[4],
        "coefficient_czk_mwh": summary_row[5] if is_export else None,
        "headers": headers,
        "rows": rows,
    }


def _validate_document_signature(extension: str, data: bytes) -> None:
    if extension == ".pdf":
        if not data.startswith(b"%PDF-"):
            raise ValueError("Soubor nemá platný PDF podpis.")
        return
    if not data.startswith(b"PK\x03\x04"):
        raise ValueError("Soubor nemá platný XLSX podpis.")
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as archive:
            members = archive.infolist()
            if not any(item.filename == "[Content_Types].xml" for item in members):
                raise ValueError("Soubor není platný XLSX dokument.")
            if sum(item.file_size for item in members) > MAX_XLSX_UNCOMPRESSED_SIZE:
                raise ValueError("Rozbalený XLSX dokument překračuje bezpečnostní limit.")
    except zipfile.BadZipFile as exc:
        raise ValueError("Soubor není platný XLSX dokument.") from exc


class InvoiceArchiveService:
    def __init__(self, root_dir: Path):
        self.root_dir = Path(root_dir)
        self.originals_dir = self.root_dir / "originals"
        self.normalized_dir = self.root_dir / "normalized"
        self.audits_dir = self.root_dir / "audits"
        for path in (self.originals_dir, self.normalized_dir, self.audits_dir):
            path.mkdir(parents=True, exist_ok=True)

    def list_documents(self) -> list[dict[str, Any]]:
        documents = []
        for path in sorted(self.normalized_dir.glob("*.json"), reverse=True):
            try:
                record = json.loads(path.read_text(encoding="utf-8"))
                parsed = dict(record.get("parsed", {}))
                parsed.pop("raw_text", None)
                parsed.pop("rows", None)
                documents.append({**record, "parsed": parsed})
            except (OSError, json.JSONDecodeError):
                continue
        return documents

    def get_document(self, document_id: str) -> dict[str, Any] | None:
        path = self.normalized_dir / f"{document_id}.json"
        if not path.exists():
            return None
        try:
            return json.loads(path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            return None

    def store(self, filename: str, data: bytes) -> dict[str, Any]:
        extension = Path(filename).suffix.lower()
        if extension not in ALLOWED_EXTENSIONS:
            raise ValueError("Podporovány jsou pouze PDF a XLSX soubory.")
        if not data or len(data) > MAX_DOCUMENT_SIZE:
            raise ValueError("Dokument je prázdný nebo překračuje limit 25 MB.")
        _validate_document_signature(extension, data)
        digest = hashlib.sha256(data).hexdigest()
        document_id = digest[:20]
        safe_name = re.sub(r"[^A-Za-z0-9._-]+", "_", Path(filename).name)
        parsed = parse_invoice_pdf(data) if extension == ".pdf" else parse_invoice_xlsx(data)
        original_path = self.originals_dir / f"{document_id}-{safe_name}"
        original_path.write_bytes(data)
        record = {
            "id": document_id,
            "filename": safe_name,
            "sha256": digest,
            "size_bytes": len(data),
            "uploaded_at": datetime.now(timezone.utc).isoformat().replace("+00:00", "Z"),
            "parsed": parsed,
        }
        (self.normalized_dir / f"{document_id}.json").write_text(json.dumps(record, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
        return record

    def delete(self, document_id: str) -> bool:
        normalized = self.normalized_dir / f"{document_id}.json"
        if not normalized.exists():
            return False
        try:
            record = json.loads(normalized.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            record = {}
        for path in self.originals_dir.glob(f"{document_id}-*"):
            path.unlink(missing_ok=True)
        normalized.unlink(missing_ok=True)
        shutil.rmtree(self.audits_dir / document_id, ignore_errors=True)
        return True

    def audit(self, document_id: str, virtual_invoice: dict[str, Any]) -> dict[str, Any]:
        record = self.get_document(document_id)
        if record is None:
            raise FileNotFoundError(document_id)
        reference = record.get("parsed", {})
        invoice = virtual_invoice.get("invoice", {}).get("actual", {})
        document_type = reference.get("document_type")
        if document_type == "invoice_pdf":
            mappings = {
                "supply_without_vat": invoice.get("supply_without_vat"),
                "commercial_without_vat": invoice.get("commercial", {}).get("total"),
                "regulated_without_vat": invoice.get("regulated", {}).get("total"),
                "export_total": invoice.get("sell_total"),
            }
        elif document_type == "supply_detail_xlsx":
            mappings = {
                "quantity_mwh": (virtual_invoice.get("actual", {}).get("kwh_total") or 0.0) / 1000.0,
                "total_czk": invoice.get("commercial", {}).get("spot_energy"),
            }
        else:
            mappings = {
                "quantity_mwh": (virtual_invoice.get("actual", {}).get("export_kwh_total") or 0.0) / 1000.0,
                "total_czk": invoice.get("sell_total"),
            }
        comparisons = []
        for key, actual in mappings.items():
            expected = reference.get(key)
            if expected is None or actual is None:
                continue
            difference = round(float(actual) - float(expected), 6)
            absolute = abs(difference)
            level = "match" if absolute <= 0.02 else "warning" if absolute <= 1.0 else "error"
            comparisons.append({"field": key, "expected": expected, "actual": actual, "difference": difference, "level": level})
        overall = "error" if any(item["level"] == "error" for item in comparisons) else "warning" if any(item["level"] == "warning" for item in comparisons) else "match"
        result = {"document_id": document_id, "overall": overall, "comparisons": comparisons, "audited_at": datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")}
        audit_dir = self.audits_dir / document_id
        audit_dir.mkdir(parents=True, exist_ok=True)
        (audit_dir / "latest.json").write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
        return result
